import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows



def modified_file_creation(each_file):
    print("Esta es mi funcion")

    file_name = os.path.basename(each_file).split('.')[0]

    # Open the excel file con OpenPyXL
    book = load_workbook(each_file)

    # Operations to apply in file
    worksheet = book['GL']
    print(worksheet.title)

    #Add new column 
    new_column = 'AX'
    new_column2 = 'AY'

    #Cell reference
    cell_new_column = worksheet[new_column + '1' ]
    cell_new_column2 = worksheet[new_column2 + '1' ]

    #Name new column
    new_column_name = "Aging in days"
    cell_new_column.value = new_column_name

    new_column_name2 = "Over a year"
    cell_new_column2.value = new_column_name2

    max_row = worksheet.max_row
    for row in range(2, max_row + 1):  # Assuming the data starts from the second row
    # Get the date from column X
        date_cell = worksheet.cell(row=row, column=24)  # Column X is the 24th column
        cell_value = date_cell.value
        if cell_value is not None:
            date_value = datetime.strptime(cell_value, "%m/%d/%Y")
            print(f"Type {type(date_value)}")
            print(date_value)

            if date_value is not None and isinstance(date_value, datetime):
                # Calculate the number of days between the date and 11/30/2023
                target_date = datetime(2023, 11, 30)
                print(target_date)
                days_difference = (target_date - date_value).days
                print(days_difference)

                # Update the value in column AX (50th column)
                worksheet.cell(row=row, column=50, value=days_difference)

            # Iterate through each row and check the value in column AX
    for row in range(2, max_row + 1):  # Assuming the data starts from the second row
        # Get the value from column AX
        ax_value = worksheet.cell(row=row, column=50).value  # Assuming column AX is the 50th column
        print(ax_value)

    # Check if the value is over 365
        if ax_value is not None and ax_value > 365:
            # Update the corresponding cell in column AW (assuming column AY is the 51th column)
            worksheet.cell(row=row, column=51, value='Yes')    
        if ax_value is not None and ax_value < 366:
            # Update the corresponding cell in column AW (assuming column AY is the 51th column)
            worksheet.cell(row=row, column=51, value='No')   
        if ax_value is None and ax_value:
            # Update the corresponding cell in column AW (assuming column AY is the 51th column)
            worksheet.cell(row=row, column=51, value='')   
    

    #Save file
    folderUbication2 = 'E:/Resources/NAM US/Python/QA Accrued trade payables modified'
    name_specifications = folderUbication2 + "/" + file_name + " modified" + ".xlsx"
    book.save(name_specifications)
# Cierra el libro después de trabajar con él
    book.close()

def pivot_table_creation (each_file2):
    #reopen the modify workbook 
    
    workbookModified = load_workbook(each_file2)
    workbookModified_name = os.path.basename(each_file2).split('.')[0]
    print(f"the name of the workbook is: {workbookModified_name}")

    #create a dataframe with the 

    sheet_name2='GL'
    df = pd.read_excel(each_file2, sheet_name=sheet_name2)   

    #data frame validations
    print(df.head())

    # Creating a pivot table
    pivot_table = df.pivot_table(values='Company code', index='Document type', columns='Over a year', aggfunc='count', fill_value=0)

    # Displaying the pivot table
    print(pivot_table.to_string())

    # Define the folder path and save the pivot table to an Excel file using openpyxl
    folder_path_pivot_table = 'E:/Resources/NAM US/Python/QA Accrued trade payables pivot tables'
    output_file_path = folder_path_pivot_table + "/"+workbookModified_name + '_pivot_table.xlsx'
    pivot_table.to_excel(output_file_path, engine='openpyxl', sheet_name='PivotTable')
    workbookModified.close()


# Search a file in a specific directory
folderUbication ='E:/Resources/NAM US/Python/QA Accrued trade payables'

#Get list of files in the folder
filesOnFolder = os.listdir(folderUbication)

#Print list of files
for file_to_work in filesOnFolder:
    each_file = os.path.join(folderUbication, file_to_work)
    modified_file_creation(each_file)

# Search a file in a specific directory
folderUbication_modified ='E:/Resources/NAM US/Python/QA Accrued trade payables modified'

#Get list of files in the folder
filesOnFolderModified = os.listdir(folderUbication_modified)

#Print list of files
for file_to_work2 in filesOnFolderModified:
    each_file2 = os.path.join(folderUbication_modified, file_to_work2)
    pivot_table_creation(each_file2)



'''
#filter excel files
excelFiles = [file for file in filesOnFolder if file.endswith('.xlsx')]

#Check if there are excel files in the folder
if excelFiles:
    # Get the first file 
    first_file = os.path.join(folderUbication, excelFiles[0])

    #Get the name of first file
    file_name = os.path.basename(first_file).split('.')[0]
    print(file_name)


    # Open the excel file con OpenPyXL
    book = load_workbook(first_file)

    # Operations to apply in file
    worksheet = book['GL']
    print(worksheet.title)

    #Add new column 
    new_column = 'AX'
    new_column2 = 'AY'

    #Cell reference
    cell_new_column = worksheet[new_column + '1' ]
    cell_new_column2 = worksheet[new_column2 + '1' ]

    #Name new column
    new_column_name = "Aging in days"
    cell_new_column.value = new_column_name

    new_column_name2 = "Over a year"
    cell_new_column2.value = new_column_name2

    max_row = worksheet.max_row

    for row in range(2, max_row + 1):  # Assuming the data starts from the second row
    # Get the date from column X
        date_cell = worksheet.cell(row=row, column=24)  # Column X is the 24th column
        cell_value = date_cell.value
        if cell_value is not None:
            date_value = datetime.strptime(cell_value, "%m/%d/%Y")
            print(f"Type {type(date_value)}")
            print(date_value)

            if date_value is not None and isinstance(date_value, datetime):
                # Calculate the number of days between the date and 11/30/2023
                target_date = datetime(2023, 11, 30)
                print(target_date)
                days_difference = (target_date - date_value).days
                print(days_difference)

                # Update the value in column AX (50th column)
                worksheet.cell(row=row, column=50, value=days_difference)

            # Iterate through each row and check the value in column AX
    for row in range(2, max_row + 1):  # Assuming the data starts from the second row
        # Get the value from column AX
        ax_value = worksheet.cell(row=row, column=50).value  # Assuming column AX is the 50th column
        print(ax_value)

    # Check if the value is over 365
        if ax_value is not None and ax_value > 365:
            # Update the corresponding cell in column AW (assuming column AY is the 51th column)
            worksheet.cell(row=row, column=51, value='Yes')    
        if ax_value is not None and ax_value < 366:
            # Update the corresponding cell in column AW (assuming column AY is the 51th column)
            worksheet.cell(row=row, column=51, value='No')   
        if ax_value is None and ax_value:
            # Update the corresponding cell in column AW (assuming column AY is the 51th column)
            worksheet.cell(row=row, column=51, value='')   
    

    #Save file
   
    name_specifications = folderUbication + "/" + file_name + " modified" + ".xlsx"
    book.save(name_specifications)



# Cierra el libro después de trabajar con él
    book.close()
else:
    print("There are not excel files in this folder")


#reopen the modify workbook 
print(name_specifications)
workbookModified = load_workbook(name_specifications)
workbookModified_name = os.path.basename(name_specifications)
print(f"the name of the workbook is: {workbookModified_name}")

#create a dataframe with the 

sheet_name2='GL'
df = pd.read_excel(name_specifications, sheet_name=sheet_name2)   

#data frame validations
print(df.head())

# Creating a pivot table
pivot_table = df.pivot_table(values='Company code', index='Document type', columns='Over a year', aggfunc='count', fill_value=0)

# Displaying the pivot table
print(pivot_table.to_string())

# Define the folder path and save the pivot table to an Excel file using openpyxl
output_file_path = folderUbication + "/"+file_name+ '_pivot_table.xlsx'
pivot_table.to_excel(output_file_path, engine='openpyxl', sheet_name='PivotTable')

    
'''

    
    