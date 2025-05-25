

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Border, Side

# Cargar el libro de trabajo existente
workbook = load_workbook('Conciliaciones.xlsx')
sheet1 = workbook['Conciliaciones']
sheet2 = workbook['Tiempo']


# Asumimos que la tabla comienza en la primera fila y tiene encabezados
# y que las columnas F y E son las columnas 6 y 5 respectivamente
columna_E = 5
columna_D = 5

# Extraer los datos de la tabla (excluyendo los encabezados)
datos = []
for row in sheet1.iter_rows(min_row=2, values_only=True):
    datos.append(list(row))

# Ordenar los datos primero por la columna F y luego por la columna E
datos_ordenados = sorted(datos, key=lambda x: (x[columna_E - 1], x[columna_D - 1]))

# Escribir los datos ordenados de nuevo en el archivo de Excel
for idx, row in enumerate(datos_ordenados, start=2):
    for col_num, value in enumerate(row, start=1):
        sheet1.cell(row=idx, column=col_num, value=value)

# Guardar el archivo de Excel con los datos ordenados
workbook.save('Conciliaciones_ordenadas.xlsx')
print("La tabla ha sido ordenada y guardada en 'Conciliaciones_ordenadas.xlsx'")

