import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Border, Side
from itertools import product

# Cargar el libro de trabajo existente
workbook = load_workbook('bddistribucion.xlsx')
sheetRollup = workbook['Rollup']
sheetRecs = workbook['Recs']

# Leer los datos de las columnas A y B en 'Hoja1'
columna_A = [celda.value for celda in sheetRollup['A'] if celda.value is not None]
columna_B = [celda.value for celda in sheetRollup['B'] if celda.value is not None]


# Leer los datos de la columna C en 'Hoja2'
columna_C = [celda.value for celda in sheetRecs['A'] if celda.value is not None]


# Combinar los elementos de las columnas A y B como conjuntos (pares)
pares_AB = [(a, b) for a, b in zip(columna_A, columna_B)]

# Generar las combinaciones posibles entre los pares (A, B) y cada elemento de la columna C
combinaciones = [(a, b, c) for (a, b) in pares_AB for c in columna_C]

# Crear una nueva hoja para guardar las combinaciones
hoja_combinaciones = workbook.create_sheet(title='Combinaciones')

# Escribir las combinaciones en la nueva hoja
for combinacion in combinaciones:
    hoja_combinaciones.append(combinacion)
    
# Guardar el libro de trabajo actualizado
workbook.save('bddistribuciones_conlistado.xlsx')
