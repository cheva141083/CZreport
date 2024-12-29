

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Border, Side

# Cargar el libro de trabajo existente
workbook = load_workbook('Divisor.xlsx')
sheet1 = workbook['Base']
sheet2 = workbook['Resultado 1']


# Obtener la columna a copiar (ajusta la columna y el rango según tus necesidades)
columna_a_copiar = 'C'
rango_columnas = sheet1[columna_a_copiar]

# Pegar la columna en la nueva hoja (ajusta la fila inicial según sea necesario)
fila_inicial = 1
for i, cell in enumerate(rango_columnas):
    sheet2.cell(row=fila_inicial + i, column=1, value=cell.value)

# Encontrar la columna a partir de la cual quieres separar (ajusta si es necesario)
column_to_split = "A"  # Suponiendo que los datos están en la columna A

# Obtener el número de filas
max_row = sheet2.max_row

# Crear nuevas columnas para los datos separados
sheet2.insert_cols(2)  # Inserta una columna antes de la columna C
sheet2.insert_cols(3)

# Iterar sobre las filas y separar los datos
for row in range(1, max_row + 1):  # Comenzando desde la fila 1 (suponiendo que la primera fila no es encabezado)
    cell_value = sheet2[column_to_split + str(row)].value
    if cell_value:
        split_values = cell_value.split("-")
        sheet2.cell(row=row, column=2, value=split_values[0])
        sheet2.cell(row=row, column=3, value=split_values[1])

# Crear una nueva hoja de trabajo para almacenar los datos copiados
new_worksheet = workbook.create_sheet("Datos canciones")

# Iterar sobre las filas, copiando los valores de las columnas B y C de las filas impares
row_num = 1
for row in sheet2.iter_rows(min_row=1, min_col=2, max_col=3):
    if row_num % 2 != 0:
        new_worksheet.append([cell.value for cell in row])
    row_num += 1

# Crear una nueva hoja de trabajo para almacenar los datos copiados
new_worksheet2 = workbook.create_sheet("Puntos")

# Iterar sobre las filas, copiando los valores de las columnas B y C de las filas impares
row_num = 1
for row in sheet2.iter_rows(min_row=1, min_col=3, max_col=3):
    if row_num % 2 == 0:
        new_worksheet2.append([cell.value for cell in row])
    row_num += 1

workbook.save('DivisorEdited.xlsx')
