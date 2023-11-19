import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Border, Side

# Cargar el libro de trabajo existente
workbook = load_workbook('Copypega.xlsx')
sheet = workbook['Mensual']

# Crear un estilo con el formato deseado (dos decimales)
currency_style = NamedStyle(name='currency',number_format='#,##0.00')

# Obtener el rango de celdas para la columna B
columna_b = sheet['C']

# Aplicar el estilo a todas las celdas de la columna B
for cell in columna_b:
    cell.style = currency_style

# Obtener el rango de celdas para la columna D
columna_d = sheet['E']

# Aplicar el estilo a todas las celdas de la columna D
for cell in columna_d:
    cell.style = currency_style


#delete rows 

rows5 = list(sheet.iter_rows())



for row3 in reversed(rows5):
    cell_C = row3[2]
    cell_E = row3[4]
        # Check if both columns C and E have zero values
    if cell_C.value ==0 and cell_E.value==0:
            # Delete the entire row
        sheet.delete_rows(row3[0].row)




#Aply border style in column B and D

for row2 in sheet.iter_rows():
    cell_A = row2[0]
    cell_C = sheet.cell(row=cell_A.row, column=3)
    cell_E = sheet.cell(row=cell_A.row, column=5)

    
    if cell_A.value is not None:
        text_total= "Total"
        text_subtotal="Subtotal"
        if text_total in cell_A.value or text_subtotal in cell_A.value:
            totalborder = Side(style='medium', color='000000')
            cell_C.border = Border(top=totalborder)
            cell_E.border = Border(top=totalborder)

# Guardar el libro de trabajo actualizado
workbook.save('Copypegaedited.xlsx')
