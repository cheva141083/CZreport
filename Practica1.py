import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#read excel fil
excel_file = pd.read_excel('supermarket_sales.xlsx')

print(excel_file)


#create pivot table
pivot_table = excel_file.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)



#export to excel, in the report sheet, starting in the row 4

pivot_table.to_excel('sales_2021.xlsx', startrow=4, sheet_name='Report')

wb = load_workbook('sales_2021.xlsx')
work_book = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_rowValue = wb.active.min_row
max_rowValue = wb.active.max_row

#create graphic
barchart = BarChart()

data = Reference(work_book, min_col=min_column+1, max_col=max_column, min_row=min_rowValue, max_row= max_rowValue)
categorias = Reference(work_book, min_col=min_column, max_col=min_column, min_row=min_rowValue+1, max_row= max_rowValue)

barchart.add_data(data, titles_from_data = True)
barchart.set_categories(categorias)
work_book.add_chart(barchart,'B12')
barchart.title ='Sales'
barchart.style = 5

wb.save('sales_2021.xlsx')

